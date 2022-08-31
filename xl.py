from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter




class XlDataBase():
    def __init__(self) -> None:
       self.wb = load_workbook("userdata.xlsx")
       self.ws = self.wb.active

    def set_data(self, fullname:str, password:str):
        self.id = self.ws.max_row
        self.username = "User"+ str(self.id)
        self.ws['A'+str(self.id+1)] = fullname
        self.ws['B'+str(self.id+1)] = self.username
        self.ws['C'+str(self.id+1)] = password
        self.wb.save("userdata.xlsx")
        return self.username

    def get_data(self) -> list:
        d = {
            '1':'FullName',
            '2':'UserName',
            '3':'Password'
        }
        result = []
        for row in range(2, self.ws.max_row+1):
            temp, count = {}, 1
            for col in range(1, self.ws.max_column+1):
                char = get_column_letter(col)
                temp[d[str(count)]] = self.ws[char+str(row)].value
                count += 1
            result.append(temp)
        return result







# # wb = Workbook()
# wb = load_workbook("User Database.xlsx")
# worksheet = wb.active

# for row in range(1, 14):

#     for col in range(1, 3):

#         char = get_column_letter(col)

#         print(worksheet[char + str(row)].value, end = " | ")
    
#     print()








# # worksheet.title = "User Details"
# # worksheet.append(['Name', 'Password'])

# # wb.save("User Database.xlsx")
