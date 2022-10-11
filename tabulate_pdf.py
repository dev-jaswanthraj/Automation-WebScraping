import re
from sys import exec_prefix
from turtle import pd
import pdfplumber
from openpyxl import load_workbook, Workbook
from pdfplumber.page import Page
from tabula.io import read_pdf
import fitz
from fitz.utils import get_text

# class pdftoexcel:

#     def __init__(self) -> None:
#         self.__path = "C:/Users/HP/Downloads/STUCOR_TT_ND21.pdf"


#     def get_data_from_pdf(self):
#         with pdfplumber.open(self.__path) as pdf:
#             __first_page = pdf.pages[0]
#             __extracted_tabel = __first_page.extract_table()
#             # print(__extracted_tabel)
#             # return 
#             self.__Employees_social_security_number = __extracted_tabel[0][2].split("\n")
#             self.__ein = __extracted_tabel[1][0].split("\n")
#             self.__full_address = __extracted_tabel[2][0].split("\n")
#             self.__1 = __extracted_tabel[1][7].split("\n")
#             self.__2 = __extracted_tabel[1][11].split("\n")
#             self.__3 = __extracted_tabel[2][7].split("\n")
#             self.__4 = __extracted_tabel[2][11].split("\n")
#             self.__5 = __extracted_tabel[3][7].split("\n")
#             self.__6 = __extracted_tabel[3][11].split("\n")
#             self.__7 = __extracted_tabel[4][7].split("\n")
#             self.__8 = __extracted_tabel[4][11].split("\n")
#             #self.__9 = __extracted_tabel[5][11].split("\n")
#             self.__10 = __extracted_tabel[5][11].split("\n")
#             self.__control_number = __extracted_tabel[5][0].split("\n")
#             self.__first_name_initial = __extracted_tabel[6][0].split("\n")
#             self.__last_name = __extracted_tabel[6][3].split("\n")
#             self.__suff = __extracted_tabel[6][6].split("\n")
#             self.__11 = __extracted_tabel[6][7].split("\n")
#             self.__12a = __extracted_tabel[6][11].split("\n")
#             self.__12b = __extracted_tabel[7]

#             def exists(a):
#                 try:
#                     a[1]
#                     return a
#                 except:
#                     return None

#             def code12_exists(a):
#                 try:
#                     return a[-2]
#                 except:
#                     return None
                

#             return {
#                 self.__Employees_social_security_number[0]:self.__Employees_social_security_number[1], 
#                 self.__ein[0]: self.__ein[1], 
#                 self.__full_address[0]: self.__full_address[1:],
#                 self.__1[0]:exists(self.__1)[1],
#                 self.__2[0]:exists(self.__2)[1],
#                 self.__3[0]:exists(self.__3)[1],
#                 self.__4[0]:exists(self.__4)[1],
#                 self.__5[0]:exists(self.__5)[1],
#                 self.__6[0]:exists(self.__6)[1],
#                 self.__7[0]:exists(self.__7)[1],
#                 self.__8[0]:exists(self.__8)[1],
#                 self.__10[0]:exists(self.__10)[1],
#                 self.__control_number[0]: exists(self.__control_number)[1],
#                 self.__first_name_initial[0]:self.__first_name_initial[1],
#                 self.__last_name[0]:exists(self.__last_name)[1],
#                 self.__suff[0]: exists(self.__suff)[1],
#                 self.__11[0]:exists(self.__11)[1],
#                 self.__12a[0]: code12_exists(self.__12a),



#             }


        
#     def get_excel(self):

#         try:
#             self.wb = load_workbook("w-2 details of user.xlsx")
#             self.sheet1 = self.wb.active
#         except:
#             self.wb = Workbook()
#             self.sheet1 = self.wb.active
#             self.sheet1.append(list(map(str,self.__data.keys())))
#             self.wb.save("w-2 details of user.xlsx")


#         return self.sheet1


#     def add_add(self):
        
        
#         self.__data = self.get_data_from_pdf()
#         __ws = self.get_excel()


#         __ws.append(list(map(str, self.__data.values())))

#         self.wb.save("w-2 details of user.xlsx")

# obj = pdftoexcel()
# obj.get_data_from_pdf()

#data = read_pdf("C:/Users/HP/Downloads/fw2-2.pdf-edited.pdf", pages="all")
#data = fitz.open("C:/Users/HP/Downloads/fw2-2.pdf-edited.pdf")
data = fitz.open("C:/Users/HP/Downloads/fw2-2.pdf")

print(get_text(data[0]).split('\n'))#[67:])
